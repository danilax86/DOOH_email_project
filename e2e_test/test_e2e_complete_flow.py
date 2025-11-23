import os
import time
import pytest
from playwright.sync_api import expect, Page
from openpyxl import load_workbook

# Testing complete flow: login --> choose template --> fill email body --> fill contacts --> send emails

@pytest.mark.e2e
class TestCompleteEmailFlow:
    def test_login(self, page: Page, base_url: str, test_user: dict):
        page.goto(f"{base_url}/login", wait_until="domcontentloaded", timeout=30000)
        page.wait_for_selector('form', state='visible', timeout=15000)
        display_name = "Test User"
        page.fill('input[name="display_name"]', display_name)
        page.fill('input[name="email"]', test_user["username"])
        page.fill('input[name="password"]', test_user["password"])
        with page.expect_navigation(timeout=15000):
            page.click('button[type="submit"]')
        page.wait_for_selector('.template-tabs', state='visible', timeout=15000)

    def test_complete_flow_login_template_body_contacts_send(
        self, page: Page, base_url: str, test_user: dict, mailhog_client
    ):
        mailhog_client.delete_all_messages()
        excel_path = os.path.join(
            os.path.dirname(__file__), '..', 'app', 'data', 'contacts.xlsx'
        )
        assert os.path.exists(excel_path), f"Excel file not found at {excel_path}"
        wb = load_workbook(filename=excel_path, read_only=True)
        ws = wb.active
        headers = [str(cell.value).lower() for cell in ws[1]]
        email_col = headers.index('email') + 1
        expected_recipients = [
            str(row[email_col - 1]).strip()
            for row in ws.iter_rows(min_row=2, values_only=True)
            if row[email_col - 1] and str(row[email_col - 1]).strip()
        ]
        wb.close()
        assert len(expected_recipients) > 0, "No recipients found in Excel file"
        self.test_login(page, base_url, test_user)
        template_tab = page.locator('div.template-tab[data-value="check"]')
        template_tab.click()
        page.wait_for_timeout(500)
        active_tab = page.locator('div.template-tab[data-value="check"].active')
        expect(active_tab).to_be_visible()
        template_content = page.locator('#message_template').input_value()
        assert template_content.strip() != "", "Template content should not be empty"
        file_input = page.locator('input[type="file"][name="contacts_file"]')
        file_input.wait_for(state='visible', timeout=10000)
        file_input.set_input_files(excel_path)
        page.wait_for_timeout(2000)
        preview = page.locator('#preview')
        expect(preview).to_be_visible()
        preview_content = preview.text_content()
        assert preview_content is not None, "Preview should be visible"
        brand_input = page.locator('input[name="brand"]')
        brand_input.wait_for(state='visible', timeout=5000)
        brand_input.fill('Test Brand')
        period_input = page.locator('input[name="period"]')
        period_input.wait_for(state='visible', timeout=5000)
        period_input.fill('01.01.2025 - 31.12.2025')
        submit_button = page.locator('#delayed-submit')
        submit_button.wait_for(state='visible', timeout=10000)
        countdown_box = page.locator('#countdown-box')
        with page.expect_response(
            lambda response: '/send-emails' in response.url and response.request.method == 'POST',
            timeout=60000
        ) as response_info:
            submit_button.click()
            countdown_box.wait_for(state='visible', timeout=5000)
            page.wait_for_timeout(6000)
        response = response_info.value
        assert response.ok, f"Request failed with status {response.status}"
        max_wait_time = 120
        start_time = time.time()
        last_status = ""
        while time.time() - start_time < max_wait_time:
            elapsed = int(time.time() - start_time)
            status_box = page.locator("#status-box")
            status = page.locator("#status")
            status_text = ""
            if status_box.is_visible():
                status_text = status_box.text_content().strip()
            elif status.is_visible():
                status_text = status.text_content().strip()
            if status_text and status_text != last_status:
                print(f"[{elapsed:3d}s] Status: {status_text}")
                last_status = status_text
            if status_text and any(
                keyword in status_text.lower()
                for keyword in ["отправлены", "success", "завершено", "успешно"]
            ):
                print(f"[{elapsed:3d}s] Email sending completed successfully")
                break
            if status_text and any(
                keyword in status_text.lower()
                for keyword in ["ошибка", "error", "не удалось"]
            ):
                page.screenshot(path='/app/test-results/e2e_error.png')
                raise Exception(f"Error in email sending: {status_text}")
            time.sleep(1)
        else:
            page.screenshot(path='/app/test-results/e2e_timeout.png')
            raise Exception(
                f"Timed out waiting for email sending to complete. Last status: {last_status}"
            )
        time.sleep(5)
        emails = mailhog_client.get_emails()
        assert len(emails) >= len(expected_recipients), \
            f"Expected at least {len(expected_recipients)} emails, but got {len(emails)}"
        sent_recipients = []
        for email in emails:
            sent_recipients.extend(email.to)
        missing_recipients = [
            r for r in expected_recipients
            if r not in sent_recipients
        ]
        assert len(missing_recipients) == 0, \
            f"No email sent to the following recipients: {', '.join(missing_recipients)}"
        for email in emails:
            assert email.subject != "", "Email subject should not be empty"
            assert email.body != "", "Email body should not be empty"
        page.screenshot(path='/app/test-results/e2e_end.png')